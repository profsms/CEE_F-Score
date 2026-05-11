"""
Piotroski F-Score Pipeline — Visegrad 3 (Poland, Czech Republic, Hungary)
==========================================================================
Loads production Excel files, cleans data, computes full F-Score panel,
and writes corrected output Excel files with consistent formatting.

Skips Slovakia (BSSE) due to insufficient data:
  - Biotika (BSL):  no CFO, sale, or cogs
  - Tatra banka (TAT): bank — gross margin / asset turnover not applicable
  - GEVORKYAN (GVR): only 4 years, no cogs
  - PTH: no data at all

Authors: Stanisław Halkiewicz, Orjika Blessing, Dmitrii Verdun
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ── Config ────────────────────────────────────────────────────────────────────

INPUT_FILES = {
    'WSE': './data/wse_piotroski_production.xlsx',
    'PSE': './data/pse_piotroski_production.xlsx',
    'BSE': './data/bse_piotroski_production.xlsx',
}

OUTPUT_FILES = {
    'WSE': './data/wse_piotroski_production.xlsx',
    'PSE': './data/pse_piotroski_production.xlsx',
    'BSE': './data/bse_piotroski_production.xlsx',
}

TITLE_LABELS = {
    'WSE': 'WSE Piotroski F-Score — Full Panel (Visegrad Study)',
    'PSE': 'PSE Piotroski F-Score — Full Panel (Visegrad Study)',
    'BSE': 'BSE Piotroski F-Score — Full Panel (Visegrad Study)',
}

# BSE ticker cleaning map (fix "A / B" style duplicates → clean form)
BSE_TICKER_FIXES = {
    'G Richter / RICHTER':     'RICHTER',
    'AUTOWALLIS.BUD / AUTOW.BUD': 'AUTOW',
    'HUMN / HUN MINING':       'DANU',   # Danubius Hotels correct BSE ticker
}

# ── Load ──────────────────────────────────────────────────────────────────────

# Map display column headers (from output files) back to code names
DISPLAY_TO_CODE = {
    'ticker': 'ticker', 'company name': 'company_name', 'sector': 'sector',
    'country': 'country', 'exchange': 'exchange', 'listed': 'listing_year',
    'listing_year': 'listing_year', 'status': 'status', 'delisted': 'delist_year',
    'delist_year': 'delist_year', 'year': 'year',
    'net income': 'ni', 'ni': 'ni',
    'total assets': 'at', 'at': 'at',
    'cfo': 'cfo',
    'lt debt': 'lt', 'lt': 'lt',
    'curr assets': 'act', 'act': 'act',
    'curr liab': 'lct', 'lct': 'lct',
    'revenue': 'sale', 'sale': 'sale',
    'cogs': 'cogs',
    'shares out.': 'csho', 'csho': 'csho',
    'roa': 'roa', 'cfo/ta': 'cfo_ta', 'cfo_ta': 'cfo_ta',
    'leverage': 'lev', 'lev': 'lev',
    'curr ratio': 'curr_ratio', 'curr_ratio': 'curr_ratio',
    'gross margin': 'gross_margin', 'gross_margin': 'gross_margin',
    'asset turn.': 'asset_turn', 'asset_turn': 'asset_turn',
    'f_roa': 'F_ROA', 'f_droa': 'F_DROA', 'f_cfo': 'F_CFO',
    'f_accrual': 'F_ACCRUAL', 'f_dlever': 'F_DLEVER', 'f_dliquid': 'F_DLIQUID',
    'f_eq_off': 'F_EQ_OFFER', 'f_eq_offer': 'F_EQ_OFFER',
    'f_margin': 'F_MARGIN', 'f_turn': 'F_TURN',
    'n_sig': 'N_SIGNALS', 'n_signals': 'N_SIGNALS',
    'fscore': 'FSCORE',
    'fwd return': 'ret', 'ret': 'ret',
    'delist ret': 'dlret', 'dlret': 'dlret',
}

def load_panel(path: str) -> pd.DataFrame:
    """Load Panel Data sheet, robust to both original and output file formats."""
    raw = pd.read_excel(path, sheet_name='Panel Data', header=None, dtype=str)

    # Find the header row: scan for any row where 'ticker' appears (any case)
    header_idx = None
    for i, row in raw.iterrows():
        for cell in row:
            if cell is not None and str(cell).strip().lower() == 'ticker':
                header_idx = i
                break
        if header_idx is not None:
            break

    if header_idx is None:
        print(f"  WARNING: Could not find header row in {path}")
        print(f"  First 3 rows: {[list(raw.iloc[i])[:5] for i in range(min(3,len(raw)))]}")
        raise ValueError(f"No header row found in {path}")

    # Build column name list, normalising display names → code names
    raw_headers = [str(v).strip() if v not in (None, 'nan') else ''
                   for v in raw.iloc[header_idx]]
    headers = [DISPLAY_TO_CODE.get(h.lower(), h.lower().replace(' ', '_'))
               for h in raw_headers]

    # Extract data rows
    data_rows = raw.iloc[header_idx + 1:].reset_index(drop=True)
    data_rows.columns = headers
    df = data_rows.dropna(how='all').reset_index(drop=True)

    # Coerce numeric columns
    numeric_cols = ['listing_year', 'delist_year', 'year',
                    'ni', 'at', 'cfo', 'lt', 'act', 'lct',
                    'sale', 'cogs', 'csho']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    return df


# ── Clean ─────────────────────────────────────────────────────────────────────

def clean_bse_tickers(df: pd.DataFrame) -> pd.DataFrame:
    """Standardise messy BSE ticker names."""
    df['ticker'] = df['ticker'].replace(BSE_TICKER_FIXES)
    return df


def validate_data(df: pd.DataFrame, exchange: str) -> None:
    """Print a data coverage report."""
    print(f"\n{'='*60}")
    print(f"  {exchange} — Data Coverage Report")
    print(f"{'='*60}")
    core = ['ni', 'at', 'cfo', 'lt', 'act', 'lct', 'sale', 'cogs', 'csho']
    n = len(df)
    for col in core:
        if col in df.columns:
            filled = df[col].notna().sum()
            print(f"  {col:8s}: {filled:4d}/{n} ({100*filled/n:5.1f}%)")
    firms = df.groupby('ticker')['status'].first()
    active   = (firms == 'active').sum()
    delisted = (firms == 'delisted').sum()
    print(f"  Firms: {len(firms)} total  ({active} active, {delisted} delisted)")
    yrs = sorted(df['year'].dropna().unique().astype(int))
    print(f"  Years: {min(yrs)} – {max(yrs)}")


# ── F-Score Computation ───────────────────────────────────────────────────────

def compute_fscore(df: pd.DataFrame) -> pd.DataFrame:
    """
    Compute all 9 Piotroski F-Score signals plus composite score.

    Signals:
      Profitability (4):
        F_ROA      = 1 if ROA > 0
        F_DROA     = 1 if ΔROA > 0 (improvement year-on-year)
        F_CFO      = 1 if CFO/Assets > 0
        F_ACCRUAL  = 1 if CFO/Assets > ROA  (quality of earnings)

      Leverage / Liquidity (3):
        F_DLEVER   = 1 if leverage ratio declined
        F_DLIQUID  = 1 if current ratio improved
        F_EQ_OFFER = 1 if shares outstanding did NOT increase (no dilution)

      Operating Efficiency (2):
        F_MARGIN   = 1 if gross margin improved
        F_TURN     = 1 if asset turnover improved

    N_SIGNALS = count of non-missing signals used
    FSCORE    = sum of non-missing signals (NaN signals excluded from sum)
    """
    df = df.copy()
    df = df.sort_values(['ticker', 'year']).reset_index(drop=True)
    g = df.groupby('ticker')

    # ── Derived ratios ────────────────────────────────────────────────────────
    df['roa']          = df['ni']  / df['at']
    df['cfo_ta']       = df['cfo'] / df['at']
    df['lev']          = df['lt']  / df['at']
    df['curr_ratio']   = df['act'] / df['lct']
    df['gross_margin'] = np.where(
        df['sale'].notna() & df['cogs'].notna() & (df['sale'] != 0),
        (df['sale'] - df['cogs']) / df['sale'],
        np.nan
    )
    df['asset_turn']   = np.where(
        df['sale'].notna() & df['at'].notna() & (df['at'] != 0),
        df['sale'] / df['at'],
        np.nan
    )

    # ── Year-over-year lagged values (within firm) ────────────────────────────
    df['roa_lag']          = g['roa'].shift(1)
    df['lev_lag']          = g['lev'].shift(1)
    df['curr_ratio_lag']   = g['curr_ratio'].shift(1)
    df['gross_margin_lag'] = g['gross_margin'].shift(1)
    df['asset_turn_lag']   = g['asset_turn'].shift(1)
    df['csho_lag']         = g['csho'].shift(1)

    # ── Signal computation ────────────────────────────────────────────────────

    # Profitability
    df['F_ROA']     = np.where(df['roa'].notna(),
                               (df['roa'] > 0).astype(float), np.nan)

    df['F_DROA']    = np.where(df['roa'].notna() & df['roa_lag'].notna(),
                               (df['roa'] > df['roa_lag']).astype(float), np.nan)

    df['F_CFO']     = np.where(df['cfo_ta'].notna(),
                               (df['cfo_ta'] > 0).astype(float), np.nan)

    df['F_ACCRUAL'] = np.where(df['cfo_ta'].notna() & df['roa'].notna(),
                               (df['cfo_ta'] > df['roa']).astype(float), np.nan)

    # Leverage / Liquidity
    df['F_DLEVER']   = np.where(df['lev'].notna() & df['lev_lag'].notna(),
                                (df['lev'] < df['lev_lag']).astype(float), np.nan)

    df['F_DLIQUID']  = np.where(df['curr_ratio'].notna() & df['curr_ratio_lag'].notna(),
                                (df['curr_ratio'] > df['curr_ratio_lag']).astype(float), np.nan)

    df['F_EQ_OFFER'] = np.where(df['csho'].notna() & df['csho_lag'].notna(),
                                (df['csho'] <= df['csho_lag']).astype(float), np.nan)

    # Efficiency
    df['F_MARGIN']   = np.where(df['gross_margin'].notna() & df['gross_margin_lag'].notna(),
                                (df['gross_margin'] > df['gross_margin_lag']).astype(float), np.nan)

    df['F_TURN']     = np.where(df['asset_turn'].notna() & df['asset_turn_lag'].notna(),
                                (df['asset_turn'] > df['asset_turn_lag']).astype(float), np.nan)

    # ── Composite score ───────────────────────────────────────────────────────
    signal_cols = ['F_ROA', 'F_DROA', 'F_CFO', 'F_ACCRUAL',
                   'F_DLEVER', 'F_DLIQUID', 'F_EQ_OFFER',
                   'F_MARGIN', 'F_TURN']

    df['N_SIGNALS'] = df[signal_cols].notna().sum(axis=1).astype(int)
    df['FSCORE']    = df[signal_cols].sum(axis=1, min_count=1)  # NaN if all NaN

    # ── Drop helper lag columns ───────────────────────────────────────────────
    drop_cols = ['roa_lag', 'lev_lag', 'curr_ratio_lag',
                 'gross_margin_lag', 'asset_turn_lag', 'csho_lag']
    df.drop(columns=drop_cols, inplace=True)

    return df


# ── F-Score Diagnostics ───────────────────────────────────────────────────────

def print_fscore_summary(df: pd.DataFrame, exchange: str) -> None:
    computed = df[df['FSCORE'].notna()]
    print(f"\n{exchange} — F-Score Summary")
    print(f"  Rows with FSCORE: {len(computed)} / {len(df)}")
    if len(computed) > 0:
        dist = computed['FSCORE'].value_counts().sort_index()
        print(f"  Distribution: {dist.to_dict()}")
        high = (computed['FSCORE'] >= 8).sum()
        low  = (computed['FSCORE'] <= 2).sum()
        print(f"  High (≥8): {high}  |  Low (≤2): {low}")
        # Signal coverage
        signal_cols = ['F_ROA','F_DROA','F_CFO','F_ACCRUAL',
                       'F_DLEVER','F_DLIQUID','F_EQ_OFFER','F_MARGIN','F_TURN']
        print("  Signal coverage (% of total rows):")
        for s in signal_cols:
            pct = 100 * df[s].notna().sum() / len(df)
            print(f"    {s:12s}: {pct:5.1f}%")


# ── Excel Output ──────────────────────────────────────────────────────────────

# Colours matching the original WSE file style
HEADER_FILL  = PatternFill("solid", fgColor="2E2466")   # dark purple
HEADER2_FILL = PatternFill("solid", fgColor="4B3F8C")   # mid purple
ALT_FILL     = PatternFill("solid", fgColor="ECEAF5")   # light purple
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
COL_FONT     = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
DATA_FONT    = Font(name="Calibri", size=10)

THIN = Side(border_style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Columns to output and their display widths
OUTPUT_COLUMNS = [
    ('ticker',       12, 'Ticker'),
    ('company_name', 28, 'Company Name'),
    ('sector',       22, 'Sector'),
    ('country',      14, 'Country'),
    ('exchange',     10, 'Exchange'),
    ('listing_year',  9, 'Listed'),
    ('status',       10, 'Status'),
    ('delist_year',   9, 'Delisted'),
    ('year',          6, 'Year'),
    # Raw financials
    ('ni',           16, 'Net Income'),
    ('at',           16, 'Total Assets'),
    ('cfo',          16, 'CFO'),
    ('lt',           16, 'LT Debt'),
    ('act',          16, 'Curr Assets'),
    ('lct',          16, 'Curr Liab'),
    ('sale',         16, 'Revenue'),
    ('cogs',         16, 'COGS'),
    ('csho',         14, 'Shares Out.'),
    # Derived ratios
    ('roa',          12, 'ROA'),
    ('cfo_ta',       12, 'CFO/TA'),
    ('lev',          12, 'Leverage'),
    ('curr_ratio',   12, 'Curr Ratio'),
    ('gross_margin', 12, 'Gross Margin'),
    ('asset_turn',   12, 'Asset Turn.'),
    # F-Score signals
    ('F_ROA',        8, 'F_ROA'),
    ('F_DROA',       8, 'F_DROA'),
    ('F_CFO',        8, 'F_CFO'),
    ('F_ACCRUAL',    9, 'F_ACCRUAL'),
    ('F_DLEVER',     9, 'F_DLEVER'),
    ('F_DLIQUID',    9, 'F_DLIQUID'),
    ('F_EQ_OFFER',  10, 'F_EQ_OFF'),
    ('F_MARGIN',     9, 'F_MARGIN'),
    ('F_TURN',       8, 'F_TURN'),
    # Composite
    ('N_SIGNALS',    9, 'N_SIG'),
    ('FSCORE',       8, 'FSCORE'),
    # Returns (to be filled later)
    ('ret',          10, 'Fwd Return'),
    ('dlret',        10, 'Delist Ret'),
]

# Columns that are F-Score signals (highlight differently)
SIGNAL_COLS_SET = {'F_ROA','F_DROA','F_CFO','F_ACCRUAL',
                   'F_DLEVER','F_DLIQUID','F_EQ_OFFER','F_MARGIN','F_TURN',
                   'N_SIGNALS','FSCORE'}

RATIO_COLS_SET = {'roa','cfo_ta','lev','curr_ratio','gross_margin','asset_turn'}


def write_excel(df: pd.DataFrame, path: str, title: str) -> None:
    """Write the panel DataFrame to a formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Panel Data"

    cols = [(c, w, h) for (c, w, h) in OUTPUT_COLUMNS if c in df.columns]
    n_cols = len(cols)

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal='left', vertical='center',
                                      indent=1)
    ws.row_dimensions[1].height = 22

    # ── Row 2: Column headers ─────────────────────────────────────────────────
    for ci, (col, width, header) in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=ci, value=header)
        cell.font = COL_FONT
        cell.fill = HEADER2_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                    wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 30

    # ── Rows 3+: Data ─────────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(df.iterrows(), start=3):
        fill = ALT_FILL if (ri % 2 == 1) else WHITE_FILL
        for ci, (col, _, _) in enumerate(cols, start=1):
            val = row.get(col, None)
            # Convert numpy types / NaN to None for clean Excel output
            if pd.isna(val) if not isinstance(val, str) else False:
                val = None
            elif isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val)

            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            # Formatting by column type
            if col in SIGNAL_COLS_SET:
                cell.fill = PatternFill("solid", fgColor="F0EDF8")
                cell.alignment = Alignment(horizontal='center')
                if col == 'FSCORE' and val is not None:
                    if val >= 8:
                        cell.fill = PatternFill("solid", fgColor="C6EFCE")
                        cell.font = Font(name="Calibri", size=10, bold=True, color="276221")
                    elif val <= 2:
                        cell.fill = PatternFill("solid", fgColor="FFC7CE")
                        cell.font = Font(name="Calibri", size=10, bold=True, color="9C0006")
                    else:
                        cell.fill = PatternFill("solid", fgColor="FFEB9C")
                elif col in ('F_ROA','F_DROA','F_CFO','F_ACCRUAL',
                             'F_DLEVER','F_DLIQUID','F_EQ_OFFER',
                             'F_MARGIN','F_TURN') and val is not None:
                    cell.fill = (PatternFill("solid", fgColor="E2EFDA") if val == 1
                                 else PatternFill("solid", fgColor="FCE4D6"))
            elif col in RATIO_COLS_SET:
                cell.alignment = Alignment(horizontal='right')
                if val is not None:
                    cell.number_format = '0.0000'
                cell.fill = fill
            elif col in ('ni','at','cfo','lt','act','lct','sale','cogs','csho'):
                cell.alignment = Alignment(horizontal='right')
                if val is not None:
                    cell.number_format = '#,##0'
                cell.fill = fill
            elif col in ('year','listing_year','delist_year','N_SIGNALS'):
                cell.alignment = Alignment(horizontal='center')
                cell.fill = fill
            elif col == 'status':
                cell.alignment = Alignment(horizontal='center')
                if val == 'delisted':
                    cell.fill = PatternFill("solid", fgColor="FFE7E7")
                    cell.font = Font(name="Calibri", size=10, color="9C0006")
                else:
                    cell.fill = PatternFill("solid", fgColor="E7F7E7")
                    cell.font = Font(name="Calibri", size=10, color="276221")
            else:
                cell.fill = fill

    # ── Freeze panes at row 3 col 3 ──────────────────────────────────────────
    ws.freeze_panes = 'C3'

    # ── Auto-filter on header row ─────────────────────────────────────────────
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}{len(df)+2}"

    wb.save(path)
    print(f"  → Saved: {path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    results = {}

    for exchange, input_path in INPUT_FILES.items():
        print(f"\n{'#'*60}")
        print(f"  Processing: {exchange}")
        print(f"{'#'*60}")

        # Load
        df = load_panel(input_path)
        print(f"  Loaded {len(df)} rows, {df['ticker'].nunique()} firms")

        # Exchange-specific cleaning
        if exchange == 'BSE':
            df = clean_bse_tickers(df)
            print(f"  BSE tickers cleaned")

        # Validate raw data
        validate_data(df, exchange)

        # Wipe any pre-existing computed columns (recompute from scratch)
        computed_cols = ['roa', 'cfo_ta', 'lev', 'curr_ratio', 'gross_margin',
                         'asset_turn', 'F_ROA', 'F_DROA', 'F_CFO', 'F_ACCRUAL',
                         'F_DLEVER', 'F_DLIQUID', 'F_EQ_OFFER', 'F_MARGIN',
                         'F_TURN', 'N_SIGNALS', 'FSCORE']
        for col in computed_cols:
            if col in df.columns:
                df[col] = np.nan

        # Compute F-Score
        df = compute_fscore(df)

        # Diagnostics
        print_fscore_summary(df, exchange)

        # Write output
        write_excel(df, OUTPUT_FILES[exchange], TITLE_LABELS[exchange])

        results[exchange] = df

    # ── Combined summary ──────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print("  COMBINED PANEL SUMMARY (V3: PL + CZ + HU)")
    print(f"{'='*60}")
    all_df = pd.concat(results.values(), ignore_index=True)
    print(f"  Total firm-year observations : {len(all_df)}")
    print(f"  Total firms                  : {all_df['ticker'].nunique()}")
    by_exch = all_df.groupby('exchange').agg(
        firms=('ticker', 'nunique'),
        rows=('ticker', 'count'),
        fscore_rows=('FSCORE', lambda x: x.notna().sum()),
        active=('status', lambda x: (x=='active').sum()),
        delisted=('status', lambda x: (x=='delisted').sum()),
    )
    print(f"\n{by_exch.to_string()}")

    computed = all_df[all_df['FSCORE'].notna()]
    print(f"\n  Overall F-Score distribution:")
    print(f"  {computed['FSCORE'].value_counts().sort_index().to_dict()}")
    print(f"\n  High-score firms (FSCORE ≥ 8): {(computed['FSCORE']>=8).sum()}")
    print(f"  Low-score firms  (FSCORE ≤ 2): {(computed['FSCORE']<=2).sum()}")
    print(f"\n  Note: Slovakia (BSSE) excluded — insufficient data for F-Score")
    print(f"        (Biotika: no CFO/sales; Tatra banka: bank, no COGS;")
    print(f"         GEVORKYAN: only 4 years, no COGS; PTH: no data)")


if __name__ == '__main__':
    main()

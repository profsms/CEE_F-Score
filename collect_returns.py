"""
collect_returns.py — Piotroski F-Score Return Collection
=========================================================
Downloads historical price data from Stooq for all firms in the
V3 panel (WSE/Poland, PSE/Czech Republic, BSE/Hungary) and
computes forward annual holding-period returns.

Return convention (look-ahead bias safe):
  - F-Score computed from fiscal year ending 31 Dec of year t
  - 6-month reporting lag assumed → portfolio formed 1 Jul, year t+1
  - Holding period: 1 Jul year t+1  →  30 Jun year t+2
  - So F-Score row for year t gets the return from t+1 Jul to t+2 Jun

  Example: F-Score 2018 → return from Jul 2019 to Jun 2020

For delisted firms:
  - Return measured from portfolio formation date to last available
    trading date (or delisting date), then position closed at last price
  - dlret column records this terminal partial-year return

API:
  Stooq URL: https://stooq.com/q/d/l/?s={ticker}&d1={yyyymmdd}&d2={yyyymmdd}&i=d&apikey={key}
  API key read from ./stooq_apikey.txt (same folder as script)

Usage:
  python collect_returns.py

Outputs:
  ./data/wse_piotroski_production.xlsx  (ret + dlret columns filled)
  ./data/pse_piotroski_production.xlsx
  ./data/bse_piotroski_production.xlsx
  ./data/prices/                        (raw price CSVs cached locally)

Authors: Stanisław Halkiewicz, Orjika Blessing, Dmitrii Verdun
"""

import os
import time
import requests
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import date, timedelta
import warnings
warnings.filterwarnings('ignore')
try:
    import yfinance as yf
    YFINANCE_AVAILABLE = True
except ImportError:
    YFINANCE_AVAILABLE = False
    print("  Note: yfinance not installed. Run: pip install yfinance")

# ── Config ────────────────────────────────────────────────────────────────────

DATA_DIR   = Path('./data')
PRICE_DIR  = DATA_DIR / 'prices'
PRICE_DIR.mkdir(parents=True, exist_ok=True)

APIKEY_FILE = Path('./stooq_apikey.txt')

# Stooq base URLs to try in order (mirrors Julia script fallback logic)
STOOQ_URLS = [
    "https://stooq.com/q/d/l/",
    "http://stooq.com/q/d/l/",
    "https://stooq.pl/q/d/l/",
    "http://stooq.pl/q/d/l/",
]

# Rate limiting — be polite to Stooq
REQUEST_DELAY = 1.5   # seconds between requests
MAX_RETRIES   = 3

# ── Ticker mapping: panel ticker → Stooq ticker ───────────────────────────────
# WSE tickers already have .WA — just lowercase them
# PSE (Prague):   suffix .pr
# BSE (Budapest): suffix .bu

STOOQ_TICKER_MAP = {
    # ── WSE / Poland — bare ticker, no exchange suffix (confirmed via stooq_debug)
    'ALE.WA':   'ale',
    'ATG.WA':   'atg',
    'BDX.WA':   'bdx',
    'CDR.WA':   'cdr',
    'DNP.WA':   'dnp',
    'ENG.WA':   'eng',
    'GPW.WA':   'gpw',
    'JSW.WA':   'jsw',
    'KGH.WA':   'kgh',
    'LBW.WA':   'lbw',
    'PLM.WA':   'plm',

    # ── PSE / Czech Republic — bare ticker, no exchange suffix (confirmed)
    'CZG':       'czg',
    'TABAK':     'tabak',
    'CEZ':       'cez',
    'KOFOL':     'kofol',
    'PRIUA':     'priua',
    'ORCO':      'orco',
    'AAA':       'aaa',
    'NWR':       'nwr',
    'ECM':       'ecm',
    'UNIPETROL': 'unipetrol',

    # ── BSE / Hungary — .hu suffix (confirmed: richter.hu, otp.hu, 4ig.hu work)
    'RICHTER':   'richter.hu',
    'AUTOW':     'autow.hu',
    'DHOUSE':    'dhouse.hu',    # NOTE: unconfirmed — may need manual check
    '4IG':       '4ig.hu',
    'PANNERGY':  'pannergy.hu',
    'KULCSSOFT': 'kulcssoft.hu',
    'OPG':       'opg.hu',
    'EGIS':      'egis.hu',
    'DANU':      'danu.hu',
    'BTEL':      'btel.hu',
}

# ── Yahoo Finance fallback for tickers not on Stooq ─────────────────────────
# PSE (Prague): Yahoo uses .PR suffix
# BSE (Budapest): Yahoo uses .BD suffix
YAHOO_FALLBACK_MAP = {
    'AAA':       'AAA.PR',
    'CZG':       'CZG.PR',
    'ECM':       'ECM.PR',
    'KOFOL':     'KOFOL.PR',
    'NWR':       'NWR.PR',
    'ORCO':      'ORCO.PR',
    'PRIUA':     'PRIUA.PR',
    'TABAK':     'TABAK.PR',
    'UNIPETROL': 'UNIP.PR',
    'AUTOW':     'AUTOW.BD',
    'BTEL':      'BTEL.BD',
    'DANU':      'DANU.BD',
    'DHOUSE':    'DHOUSE.BD',
    'EGIS':      'EGIS.BD',
    'KULCSSOFT': 'KULCS.BD',
    'OPG':       'OPG.BD',
}

# ── Column name mapping (handles both original and output file headers) ────────

DISPLAY_TO_CODE = {
    'ticker': 'ticker', 'company name': 'company_name', 'sector': 'sector',
    'country': 'country', 'exchange': 'exchange', 'listed': 'listing_year',
    'listing_year': 'listing_year', 'status': 'status', 'delisted': 'delist_year',
    'delist_year': 'delist_year', 'year': 'year',
    'net income': 'ni', 'ni': 'ni', 'total assets': 'at', 'at': 'at',
    'cfo': 'cfo', 'lt debt': 'lt', 'lt': 'lt',
    'curr assets': 'act', 'act': 'act', 'curr liab': 'lct', 'lct': 'lct',
    'revenue': 'sale', 'sale': 'sale', 'cogs': 'cogs',
    'shares out.': 'csho', 'csho': 'csho',
    'roa': 'roa', 'cfo/ta': 'cfo_ta', 'cfo_ta': 'cfo_ta',
    'leverage': 'lev', 'lev': 'lev',
    'curr ratio': 'curr_ratio', 'curr_ratio': 'curr_ratio',
    'gross margin': 'gross_margin', 'gross_margin': 'gross_margin',
    'asset turn.': 'asset_turn', 'asset_turn': 'asset_turn',
    'f_roa': 'F_ROA', 'f_droa': 'F_DROA', 'f_cfo': 'F_CFO',
    'f_accrual': 'F_ACCRUAL', 'f_dlever': 'F_DLEVER', 'f_dliquid': 'F_DLIQUID',
    'f_eq_off': 'F_EQ_OFFER', 'f_eq_offer': 'F_EQ_OFFER',
    'f_margin': 'F_MARGIN', 'f_turn': 'F_TURN',
    'n_sig': 'N_SIGNALS', 'n_signals': 'N_SIGNALS',
    'fscore': 'FSCORE', 'fwd return': 'ret', 'ret': 'ret',
    'delist ret': 'dlret', 'dlret': 'dlret',
}

# ── Excel output config (mirrors piotroski_pipeline.py) ───────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="2E2466")
HEADER2_FILL = PatternFill("solid", fgColor="4B3F8C")
ALT_FILL     = PatternFill("solid", fgColor="ECEAF5")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
COL_FONT     = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
DATA_FONT    = Font(name="Calibri", size=10)
THIN         = Side(border_style="thin", color="CCCCCC")
THIN_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SIGNAL_COLS  = {'F_ROA','F_DROA','F_CFO','F_ACCRUAL','F_DLEVER',
                'F_DLIQUID','F_EQ_OFFER','F_MARGIN','F_TURN','N_SIGNALS','FSCORE'}
RATIO_COLS   = {'roa','cfo_ta','lev','curr_ratio','gross_margin','asset_turn'}

OUTPUT_COLUMNS = [
    ('ticker',12,'Ticker'),('company_name',28,'Company Name'),
    ('sector',22,'Sector'),('country',14,'Country'),
    ('exchange',10,'Exchange'),('listing_year',9,'Listed'),
    ('status',10,'Status'),('delist_year',9,'Delisted'),('year',6,'Year'),
    ('ni',16,'Net Income'),('at',16,'Total Assets'),('cfo',16,'CFO'),
    ('lt',16,'LT Debt'),('act',16,'Curr Assets'),('lct',16,'Curr Liab'),
    ('sale',16,'Revenue'),('cogs',16,'COGS'),('csho',14,'Shares Out.'),
    ('roa',12,'ROA'),('cfo_ta',12,'CFO/TA'),('lev',12,'Leverage'),
    ('curr_ratio',12,'Curr Ratio'),('gross_margin',12,'Gross Margin'),
    ('asset_turn',12,'Asset Turn.'),
    ('F_ROA',8,'F_ROA'),('F_DROA',8,'F_DROA'),('F_CFO',8,'F_CFO'),
    ('F_ACCRUAL',9,'F_ACCRUAL'),('F_DLEVER',9,'F_DLEVER'),
    ('F_DLIQUID',9,'F_DLIQUID'),('F_EQ_OFFER',10,'F_EQ_OFF'),
    ('F_MARGIN',9,'F_MARGIN'),('F_TURN',8,'F_TURN'),
    ('N_SIGNALS',9,'N_SIG'),('FSCORE',8,'FSCORE'),
    ('ret',12,'Fwd Return'),('dlret',12,'Delist Ret'),
    ('ret_start',13,'Ret Start'),('ret_end',13,'Ret End'),
]

TITLE_LABELS = {
    'WSE': 'WSE Piotroski F-Score — Full Panel (Visegrad Study)',
    'PSE': 'PSE Piotroski F-Score — Full Panel (Visegrad Study)',
    'BSE': 'BSE Piotroski F-Score — Full Panel (Visegrad Study)',
}

# ── API Key ───────────────────────────────────────────────────────────────────

def load_apikey() -> str:
    key = os.environ.get('STOOQ_APIKEY', '').strip()
    if not key and APIKEY_FILE.exists():
        key = APIKEY_FILE.read_text().strip()
    if not key:
        raise FileNotFoundError(
            "No Stooq API key found.\n"
            "Place your key in ./stooq_apikey.txt or set STOOQ_APIKEY env variable.\n"
            "Get a key at: https://stooq.pl/q/d/?s=kgh&get_apikey"
        )
    return key


# ── Stooq Download ────────────────────────────────────────────────────────────

def stooq_download(stooq_ticker: str, date_start: date, date_end: date,
                   apikey: str) -> pd.DataFrame:
    """
    Download daily close prices from Stooq for a single ticker.
    Returns DataFrame with columns [date, close], sorted ascending.
    Raises on failure after MAX_RETRIES attempts.
    """
    cache_file = PRICE_DIR / f"{stooq_ticker}_{date_start:%Y%m%d}_{date_end:%Y%m%d}.csv"
    if cache_file.exists():
        df = pd.read_csv(cache_file, parse_dates=['date'])
        df['close'] = pd.to_numeric(df['close'], errors='coerce')
        return df.dropna(subset=['close']).sort_values('date').reset_index(drop=True)

    d1 = date_start.strftime('%Y%m%d')
    d2 = date_end.strftime('%Y%m%d')
    params = {'s': stooq_ticker, 'd1': d1, 'd2': d2, 'i': 'd', 'apikey': apikey}
    headers = {'User-Agent': 'Mozilla/5.0'}

    body = None
    last_error = ''
    for base_url in STOOQ_URLS:
        for attempt in range(MAX_RETRIES):
            try:
                resp = requests.get(base_url, params=params, headers=headers, timeout=20)
                text = resp.text
                if 'Uzyskaj apikey' in text or 'get_apikey' in text:
                    raise ValueError(f"Stooq API key rejected for {stooq_ticker}")
                if 'No data' in text or len(text.strip()) < 20:
                    last_error = f"No data returned ({base_url})"
                    break
                first_line = text.split('\n')[0]
                if ('Date' in first_line or 'Data' in first_line) and \
                   ('Close' in first_line or 'Zamkn' in first_line):
                    body = text
                    break
                last_error = f"Unexpected response: {first_line[:80]}"
            except requests.RequestException as e:
                last_error = str(e)
                time.sleep(REQUEST_DELAY)
        if body:
            break
        time.sleep(REQUEST_DELAY)

    if not body:
        raise ConnectionError(
            "Failed to download " + stooq_ticker + ": " + last_error
        )


    # Parse CSV — Stooq uses comma or semicolon delimiter
    delim = ';' if body.count(';') > body.count(',') else ','
    from io import StringIO
    raw = pd.read_csv(StringIO(body), sep=delim)

    # Normalise column names
    raw.columns = [c.strip() for c in raw.columns]
    date_col  = next((c for c in raw.columns if c in ('Date', 'Data')), None)
    close_col = next((c for c in raw.columns if c in ('Close', 'Zamkniecie', 'Zamkn\u0119cie')), None)
    if not date_col or not close_col:
        raise ValueError(f"Cannot find Date/Close columns in Stooq response for {stooq_ticker}. "
                         f"Columns: {list(raw.columns)}")

    df = pd.DataFrame({
        'date':  pd.to_datetime(raw[date_col], errors='coerce'),
        'close': pd.to_numeric(raw[close_col], errors='coerce'),
    }).dropna().sort_values('date').reset_index(drop=True)

    # Cache to disk
    df.to_csv(cache_file, index=False)
    print(f"    Downloaded {stooq_ticker}: {len(df)} trading days "
          f"({df['date'].min().date()} → {df['date'].max().date()})")
    return df


# ── Return Computation ────────────────────────────────────────────────────────

def holding_period_return(prices: pd.DataFrame,
                          start: date, end: date) -> float | None:
    """
    Compute buy-and-hold return between start and end dates.
    Uses the nearest available trading day on or after start,
    and the nearest available trading day on or before end.
    Returns None if insufficient data.
    """
    p = prices.copy()
    p['date'] = pd.to_datetime(p['date']).dt.date

    # Entry: first available day >= start
    entry_rows = p[p['date'] >= start]
    if entry_rows.empty:
        return None
    entry_price = entry_rows.iloc[0]['close']
    entry_date  = entry_rows.iloc[0]['date']

    # Exit: last available day <= end
    exit_rows = p[p['date'] <= end]
    if exit_rows.empty:
        return None
    exit_price = exit_rows.iloc[-1]['close']
    exit_date  = exit_rows.iloc[-1]['date']

    if entry_date >= exit_date:
        return None

    return (exit_price / entry_price) - 1.0


def compute_returns_for_firm(ticker: str, panel_rows: pd.DataFrame,
                             prices: pd.DataFrame) -> pd.DataFrame:
    """
    For each firm-year row in panel_rows, compute:
      ret   = 12-month holding return from Jul (year+1) to Jun (year+2)
      dlret = partial return to delisting date for delisted firms' final year
      ret_start / ret_end = actual dates used (for transparency)

    The 6-month lag (Jul 1 of year+1) avoids look-ahead bias:
    annual reports for fiscal year t are typically filed by Apr-Jun of t+1.
    """
    rows = panel_rows.copy().sort_values('year').reset_index(drop=True)
    status   = rows['status'].iloc[0]
    is_delisted = (status == 'delisted')
    delist_yr = rows['delist_year'].iloc[0] if 'delist_year' in rows.columns else None

    ret_vals   = []
    dlret_vals = []
    start_vals = []
    end_vals   = []

    for _, row in rows.iterrows():
        yr = int(row['year'])

        # Standard return window: Jul 1 (yr+1) → Jun 30 (yr+2)
        ret_start = date(yr + 1, 7, 1)
        ret_end   = date(yr + 2, 6, 30)

        # For delisted firms, check if holding period overlaps with delisting
        if is_delisted and delist_yr is not None:
            delist_date = date(int(delist_yr), 12, 31)  # conservative end-of-year
            # Find the actual last trading day from price data
            last_trade = None
            if prices is not None and not prices.empty:
                p_dates = pd.to_datetime(prices['date']).dt.date
                valid = p_dates[p_dates <= delist_date]
                if not valid.empty:
                    last_trade = valid.max()

            # If delist falls within holding period → truncate
            if last_trade is not None and last_trade < ret_end:
                actual_end = last_trade
                r = holding_period_return(prices, ret_start, actual_end)
                ret_vals.append(r)
                dlret_vals.append(r)   # this IS the terminal return
                start_vals.append(str(ret_start))
                end_vals.append(str(actual_end))
                continue

        # Standard case
        r = holding_period_return(prices, ret_start, ret_end) if prices is not None else None
        ret_vals.append(r)
        dlret_vals.append(None)
        start_vals.append(str(ret_start) if r is not None else None)
        end_vals.append(str(ret_end) if r is not None else None)

    rows['ret']       = ret_vals
    rows['dlret']     = dlret_vals
    rows['ret_start'] = start_vals
    rows['ret_end']   = end_vals
    return rows


# ── Panel Load ────────────────────────────────────────────────────────────────

def load_panel(path: str) -> pd.DataFrame:
    raw = pd.read_excel(path, sheet_name='Panel Data', header=None, dtype=str)
    for i, row in raw.iterrows():
        if any(str(v).strip().lower() == 'ticker' for v in row if v not in (None, 'nan')):
            headers = [
                DISPLAY_TO_CODE.get(str(v).strip().lower(),
                                    str(v).strip().lower().replace(' ', '_'))
                if v not in (None, 'nan') else f'col_{j}'
                for j, v in enumerate(raw.iloc[i])
            ]
            df = raw.iloc[i + 1:].copy()
            df.columns = headers
            df = df.dropna(how='all').reset_index(drop=True)
            num_cols = ['year', 'ni', 'at', 'cfo', 'lt', 'act', 'lct', 'sale',
                        'cogs', 'csho', 'roa', 'cfo_ta', 'lev', 'curr_ratio',
                        'gross_margin', 'asset_turn', 'F_ROA', 'F_DROA', 'F_CFO',
                        'F_ACCRUAL', 'F_DLEVER', 'F_DLIQUID', 'F_EQ_OFFER',
                        'F_MARGIN', 'F_TURN', 'N_SIGNALS', 'FSCORE',
                        'listing_year', 'delist_year']
            for c in num_cols:
                if c in df.columns:
                    df[c] = pd.to_numeric(df[c], errors='coerce')
            return df
    raise ValueError(f"No header row found in {path}")


# ── Excel Write ───────────────────────────────────────────────────────────────

def write_excel(df: pd.DataFrame, path: str, title: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Panel Data"

    cols = [(c, w, h) for (c, w, h) in OUTPUT_COLUMNS if c in df.columns]
    n_cols = len(cols)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1, value=title)
    tc.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
    tc.fill = HEADER_FILL
    tc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 22

    for ci, (col, width, header) in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=ci, value=header)
        cell.font = COL_FONT
        cell.fill = HEADER2_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 30

    for ri, (_, row) in enumerate(df.iterrows(), start=3):
        fill = ALT_FILL if (ri % 2 == 1) else WHITE_FILL
        for ci, (col, _, _) in enumerate(cols, start=1):
            val = row.get(col, None)
            if isinstance(val, float) and np.isnan(val):
                val = None
            elif isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val)

            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            if col in SIGNAL_COLS:
                cell.fill = PatternFill("solid", fgColor="F0EDF8")
                cell.alignment = Alignment(horizontal='center')
                if col == 'FSCORE' and val is not None:
                    if val >= 8:
                        cell.fill = PatternFill("solid", fgColor="C6EFCE")
                        cell.font = Font(name="Calibri", size=10, bold=True, color="276221")
                    elif val <= 2:
                        cell.fill = PatternFill("solid", fgColor="FFC7CE")
                        cell.font = Font(name="Calibri", size=10, bold=True, color="9C0006")
                    else:
                        cell.fill = PatternFill("solid", fgColor="FFEB9C")
                elif col in ('F_ROA','F_DROA','F_CFO','F_ACCRUAL','F_DLEVER',
                             'F_DLIQUID','F_EQ_OFFER','F_MARGIN','F_TURN') and val is not None:
                    cell.fill = (PatternFill("solid", fgColor="E2EFDA") if val == 1
                                 else PatternFill("solid", fgColor="FCE4D6"))
            elif col in ('ret', 'dlret'):
                cell.alignment = Alignment(horizontal='right')
                cell.fill = fill
                if val is not None:
                    cell.number_format = '0.00%'
                    if val > 0:
                        cell.font = Font(name="Calibri", size=10, color="276221")
                    elif val < -0.1:
                        cell.font = Font(name="Calibri", size=10, color="9C0006")
            elif col in RATIO_COLS:
                cell.alignment = Alignment(horizontal='right')
                if val is not None:
                    cell.number_format = '0.0000'
                cell.fill = fill
            elif col in ('ni','at','cfo','lt','act','lct','sale','cogs','csho'):
                cell.alignment = Alignment(horizontal='right')
                if val is not None:
                    cell.number_format = '#,##0'
                cell.fill = fill
            elif col in ('year','listing_year','delist_year','N_SIGNALS'):
                cell.alignment = Alignment(horizontal='center')
                cell.fill = fill
            elif col == 'status':
                cell.alignment = Alignment(horizontal='center')
                if val == 'delisted':
                    cell.fill = PatternFill("solid", fgColor="FFE7E7")
                    cell.font = Font(name="Calibri", size=10, color="9C0006")
                else:
                    cell.fill = PatternFill("solid", fgColor="E7F7E7")
                    cell.font = Font(name="Calibri", size=10, color="276221")
            else:
                cell.fill = fill

    ws.freeze_panes = 'C3'
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}{len(df)+2}"
    wb.save(path)


# ── Yahoo Finance Download ────────────────────────────────────────────────────

def yahoo_download(yahoo_ticker: str, date_start: date, date_end: date) -> pd.DataFrame:
    """Download daily close prices from Yahoo Finance as fallback."""
    cache_file = PRICE_DIR / f"yahoo_{yahoo_ticker.replace('.','_')}_{date_start:%Y%m%d}_{date_end:%Y%m%d}.csv"
    if cache_file.exists():
        df = pd.read_csv(cache_file, parse_dates=['date'])
        df['close'] = pd.to_numeric(df['close'], errors='coerce')
        return df.dropna(subset=['close']).sort_values('date').reset_index(drop=True)

    if not YFINANCE_AVAILABLE:
        raise ImportError("yfinance not installed")

    raw = yf.download(yahoo_ticker,
                      start=date_start.strftime('%Y-%m-%d'),
                      end=date_end.strftime('%Y-%m-%d'),
                      progress=False, auto_adjust=True)
    if raw.empty or len(raw) < 5:
        raise ValueError(f"Yahoo returned no data for {yahoo_ticker}")

    # yfinance returns MultiIndex columns when downloading single ticker
    if isinstance(raw.columns, pd.MultiIndex):
        raw.columns = raw.columns.get_level_values(0)

    df = pd.DataFrame({
        'date':  pd.to_datetime(raw.index),
        'close': pd.to_numeric(raw['Close'], errors='coerce'),
    }).dropna().sort_values('date').reset_index(drop=True)

    df.to_csv(cache_file, index=False)
    print(f"    [Yahoo] {yahoo_ticker}: {len(df)} trading days "
          f"({df['date'].min().date()} → {df['date'].max().date()})")
    return df


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Piotroski Return Collection — Visegrad 3")
    print("=" * 55)

    # Load API key
    try:
        apikey = load_apikey()
        print(f"  API key loaded: {apikey[:6]}{'*' * (len(apikey)-6)}")
    except FileNotFoundError as e:
        print(f"\nERROR: {e}")
        return

    files = {
        'WSE': DATA_DIR / 'wse_piotroski_production.xlsx',
        'PSE': DATA_DIR / 'pse_piotroski_production.xlsx',
        'BSE': DATA_DIR / 'bse_piotroski_production.xlsx',
    }

    for exchange, path in files.items():
        print(f"\n{'#'*55}")
        print(f"  {exchange}")
        print(f"{'#'*55}")

        df = load_panel(str(path))
        tickers = sorted(df['ticker'].dropna().unique())
        print(f"  Firms: {len(tickers)}")

        # Ensure return columns exist
        for col in ('ret', 'dlret', 'ret_start', 'ret_end'):
            if col not in df.columns:
                df[col] = np.nan

        updated_chunks = []

        for ticker in tickers:
            stooq_ticker = STOOQ_TICKER_MAP.get(ticker)
            if not stooq_ticker:
                print(f"  [{ticker}] No Stooq mapping — skipping")
                chunk = df[df['ticker'] == ticker].copy()
                updated_chunks.append(chunk)
                continue

            firm_rows = df[df['ticker'] == ticker].copy()
            years = sorted(firm_rows['year'].dropna().unique().astype(int))

            # We need prices from Jul of min(year)+1 through Jun of max(year)+2
            price_start = date(min(years) + 1, 7, 1)
            price_end   = date(max(years) + 2, 6, 30)

            print(f"\n  [{ticker}] → stooq:{stooq_ticker}  "
                  f"price window: {price_start} → {price_end}")

            try:
                prices = stooq_download(stooq_ticker, price_start, price_end, apikey)
                time.sleep(REQUEST_DELAY)
            except Exception as e:
                print(f"    Stooq failed: {e}")
                # Try Yahoo Finance fallback
                yahoo_ticker = YAHOO_FALLBACK_MAP.get(ticker)
                if yahoo_ticker:
                    try:
                        print(f"    Trying Yahoo Finance: {yahoo_ticker}")
                        prices = yahoo_download(yahoo_ticker, price_start, price_end)
                        time.sleep(0.5)
                    except Exception as ye:
                        print(f"    Yahoo also failed: {ye}")
                        prices = None
                else:
                    prices = None
                if prices is None:
                    print(f"    WARNING: No price data available for {ticker}")

            chunk = compute_returns_for_firm(ticker, firm_rows, prices)
            updated_chunks.append(chunk)

            # Print summary for this firm
            ret_filled = chunk['ret'].notna().sum()
            print(f"    Returns filled: {ret_filled}/{len(chunk)} rows")
            if ret_filled > 0:
                rets = chunk['ret'].dropna()
                print(f"    Range: {rets.min():.1%} to {rets.max():.1%}  "
                      f"mean: {rets.mean():.1%}")

        # Reassemble full panel
        df_out = pd.concat(updated_chunks, ignore_index=True)
        df_out = df_out.sort_values(['ticker', 'year']).reset_index(drop=True)

        # Write back
        write_excel(df_out, str(path), TITLE_LABELS[exchange])
        print(f"\n  → Saved: {path}")

        # Summary
        total_ret = df_out['ret'].notna().sum()
        total_rows = len(df_out)
        print(f"  Returns filled: {total_ret}/{total_rows} rows")

    print(f"\n{'='*55}")
    print("Done. Prices cached in ./data/prices/")
    print("Next step: portfolio construction and performance analysis.")


if __name__ == '__main__':
    main()

"""
Biznesradar.pl WSE Financial Data Scraper — Selenium version
=============================================================
Uses a real Chrome browser (headless) to execute JavaScript and extract
the full financial table after it loads. This is necessary because
Biznesradar renders its main data tables via JavaScript.

Setup
-----
    pip install selenium webdriver-manager requests beautifulsoup4 lxml openpyxl pandas

    Chrome must be installed on your machine (any recent version).
    webdriver-manager downloads the matching ChromeDriver automatically.

Usage
-----
    python scrape_biznesradar_selenium.py
    python scrape_biznesradar_selenium.py --tickers BDX.WA CDR.WA
    python scrape_biznesradar_selenium.py --no-cache
    python scrape_biznesradar_selenium.py --no-prices
    python scrape_biznesradar_selenium.py --debug-labels BDX.WA
    python scrape_biznesradar_selenium.py --visible   # show browser window

Outputs  ./output/
----------
    financials_raw_production.csv
    wse_piotroski_production.csv
    wse_piotroski_production.xlsx
    cache/{TICKER}.json     (delete to force re-scrape)
"""

import io
import json
import time
import logging
import argparse
import unicodedata
import re
from pathlib import Path

import pandas as pd
import numpy as np
from bs4 import BeautifulSoup

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

OUT_DIR   = Path("output")
CACHE_DIR = OUT_DIR / "cache"
OUT_DIR.mkdir(exist_ok=True)
CACHE_DIR.mkdir(exist_ok=True)

# ── Firm universe ─────────────────────────────────────────────────────
FIRMS = [
    # Active
    ("ALE.WA", "ALLEGRO",   "Allegro.eu S.A.",                                 "E-commerce",                  2017, None, "active"),
    ("ALR.WA", "ALIOR",     "Alior Bank S.A.",                                 "Banking",                     2012, None, "active"),
    ("BDX.WA", "BUDIMEX",   "Budimex S.A.",                                    "Construction",                1995, None, "active"),
    ("CCC.WA", "CCC",       "CCC S.A. (renamed Modivo S.A. / MDV in Feb 2026)", "Retail / Fashion",           2004, None, "active"),
    ("CDR.WA", "CDR",       "CD Projekt S.A.",                                 "Gaming",                      2010, None, "active"),
    ("DNP.WA", "DNP",       "Dino Polska S.A.",                                "Grocery Retail",              2017, None, "active"),
    ("ENG.WA", "ENERGA",    "Energa S.A.",                                     "Utilities",                   2013, None, "active"),
    ("GPW.WA", "GPW",       "Gielda Papierow Wartosciowych w Warszawie S.A.",  "Stock Exchange Operator",     2010, None, "active"),
    ("JSW.WA", "JSW",       "Jastrzebska Spolka Weglowa S.A.",                 "Steel / Coal",                2011, None, "active"),
    ("KGH.WA", "KGHM",      "KGHM Polska Miedz S.A.",                          "Metals & Mining",             1997, None, "active"),
    # Delisted — existing
    ("IDA.WA", "IDA",       "Idea Bank S.A.",                                  "Banking",                     2015, 2020, "delisted"),
    ("PLM.WA", "PLM",       "Play Communications S.A.",                        "Telecommunications",          2017, 2022, "delisted"),
    ("TIM.WA", "TIM",       "TIM S.A.",                                         "Electrical Wholesale",        1998, 2024, "delisted"),
    # Delisted — new additions
    ("GNB.WA", "GNB",       "Getin Noble Bank S.A.",                           "Banking",                     2004, 2022, "delisted"),
    ("BOS.WA", "BOS",       "Bank Ochrony Srodowiska S.A.",                    "Banking",                     1997, 2023, "delisted"),
    ("CIE.WA", "CIE",       "Ciech S.A.",                                      "Chemicals",                   2005, 2024, "delisted"),
    ("ATG.WA", "ATG",       "Arteria S.A.",                                    "Business Services",           2007, 2021, "delisted"),
    ("LBW.WA", "BOGDANKA",  "Lubelski Wegiel Bogdanka S.A.",                   "Coal Mining",                 2009, 2016, "delisted"),
    ("PGF.WA", "PGF",       "Polska Grupa Farmaceutyczna S.A.",                "Pharmaceuticals / Wholesale", 2001, 2012, "delisted"),
]


BANK_TICKERS = {"ALR.WA", "BOS.WA", "GNB.WA", "IDA.WA"}
DEFAULT_EXCLUDED_TICKERS = {"CCC.WA", "CIE.WA", "GNB.WA", "IDA.WA", "PGF.WA", "TIM.WA"}


REPORT_TYPES = {
    "income":   "raporty-finansowe-rachunek-zyskow-i-strat",
    "balance":  "raporty-finansowe-bilans",
    "cashflow": "raporty-finansowe-przeplywy-pieniezne",
}

BASE_URL     = "https://www.biznesradar.pl"
PAGE_TIMEOUT = 30   # seconds to wait for table to appear
DELAY        = 2.0  # seconds between page loads


# ═════════════════════════════════════════════════════════════════════
# SELENIUM DRIVER SETUP
# ═════════════════════════════════════════════════════════════════════

def make_driver(headless: bool = True):
    """
    Create a Selenium Chrome WebDriver.
    webdriver-manager handles ChromeDriver version matching automatically.
    """
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager

    opts = Options()
    if headless:
        opts.add_argument("--headless=new")          # Chrome >= 112 headless mode
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--lang=pl-PL")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
    # Suppress DevTools noise
    opts.add_experimental_option("excludeSwitches", ["enable-logging"])
    opts.add_experimental_option("prefs", {"intl.accept_languages": "pl,pl-PL"})

    service = Service(ChromeDriverManager().install())
    driver  = webdriver.Chrome(service=service, options=opts)
    driver.set_page_load_timeout(30)
    return driver


def wait_for_table(driver, timeout: int = PAGE_TIMEOUT) -> bool:
    """
    Wait until the financial data table is present and has data rows.
    Biznesradar populates the table asynchronously; we poll until rows appear.
    Returns True if table loaded, False if timed out.
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    try:
        # Wait for a table row that contains a year (column header)
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located(
                (By.XPATH, "//table//th[contains(text(),'20')]")
            )
        )
        # Extra pause for row data to populate
        time.sleep(2.5)
        return True
    except Exception:
        return False


# ═════════════════════════════════════════════════════════════════════
# TABLE PARSING  (identical logic to requests version, applied to
# the fully-rendered page source)
# ═════════════════════════════════════════════════════════════════════

# Polish characters not decomposable by NFKD — must be substituted explicitly
_PL = str.maketrans({
    'ł':'l','Ł':'L','ó':'o','Ó':'O','ę':'e','Ę':'E','ą':'a','Ą':'A',
    'ś':'s','Ś':'S','ź':'z','Ź':'Z','ż':'z','Ż':'Z','ć':'c','Ć':'C',
    'ń':'n','Ń':'N','ú':'u','Ú':'U',
})

def norm(text: str) -> str:
    t = text.translate(_PL)
    t = unicodedata.normalize("NFKD", t)
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", t.lower().strip())


def clean_number(text: str):
    if not text:
        return None
    t = (text.strip()
             .replace("\xa0","").replace("\u202f","").replace("\u00a0","")
             .replace(" ","").replace("\t","")
             .replace(",","."))
    # Handle em-dash / en-dash as zero
    if t in ("-", "–", "—", ""):
        return None
    if t.startswith("(") and t.endswith(")"):
        t = "-" + t[1:-1]
    t = re.sub(r"[%złPLN$€]", "", t)
    try:
        v = float(t)
        return v if not (pd.isna(v) or abs(v) > 1e15) else None
    except (ValueError, TypeError):
        return None




def extract_table_rows_via_js(driver):
    """
    Ask the live browser for visible table cell text. This is more reliable than
    reparsing page_source when a site hydrates table contents via JavaScript.
    Returns a list of tables, each table being a list of rows, each row a list of cell texts.
    """
    script = """
    return Array.from(document.querySelectorAll('table')).map(t =>
      Array.from(t.querySelectorAll('tr')).map(tr =>
        Array.from(tr.querySelectorAll('th,td')).map(c => (c.innerText || c.textContent || '').trim())
      )
    );
    """
    try:
        return driver.execute_script(script) or []
    except Exception as e:
        log.debug(f"  JS table extraction failed: {e}")
        return []


def parse_rows_matrix(tables, ticker: str = "") -> dict:
    """Parse the live browser-extracted table matrix into {label: {year: value}}."""
    result = {}
    main_table = None
    years = []

    for table in tables:
        for row in table:
            row_years = []
            for cell in row:
                for m in re.finditer(r"\b(20\d{2})\b", cell or ""):
                    y = int(m.group(1))
                    if 2010 <= y <= 2024:
                        row_years.append(y)
            if len(row_years) >= 2:
                main_table = table
                years = row_years
                break
        if main_table is not None:
            break

    if main_table is None or not years:
        return result

    found_header = False
    for row in main_table:
        if not row:
            continue

        row_has_year = any(re.search(r"\b20\d{2}\b", cell or "") for cell in row)
        if not found_header:
            if row_has_year:
                found_header = True
            continue

        label = norm(row[0])
        if not label or len(label) < 2:
            continue

        raw_vals = row[1:]
        if not raw_vals:
            continue

        values = {}
        # Typical visible-browser layout: one cell per year after label
        for year, cell in zip(years, raw_vals[:len(years)]):
            v = clean_number(cell)
            if v is not None:
                values[year] = v

        # If the row is messier, flatten cell text into numeric tokens and align the first N
        if len(values) <= 1:
            tokens = []
            for cell in raw_vals:
                parts = [p.strip() for p in re.split(r"[\n\r]+", cell or "") if p.strip()]
                for part in parts:
                    if re.search(r"~branza|~branża|^r/r$|^k/k$", part, re.I):
                        continue
                    if re.search(r"\b20\d{2}\b", part):
                        continue
                    v = clean_number(part)
                    if v is not None:
                        tokens.append(v)
            if len(tokens) >= 2:
                values = {year: val for year, val in zip(years, tokens[:len(years)])}

        if values:
            result[label] = values

    return result

def parse_rendered_table(html: str, ticker: str = "") -> dict:
    """
    Parse the fully JS-rendered HTML of a Biznesradar report page.
    Returns {normalised_label: {year: value}}.

    The page occasionally renders the header row with a slightly different cell
    structure than the data rows. Because of that, relying only on absolute
    header column indices can collapse a row to the first year only. We keep the
    old header-index logic as a fallback, but prefer sequential alignment of the
    data cells after the label whenever the row length matches the detected year
    count.
    """
    result = {}
    soup   = BeautifulSoup(html, "lxml")

    # Find table with year headers
    main_table = None
    for t in soup.find_all("table"):
        ths = t.find_all("th")
        if any(re.search(r"\b20(1[0-9]|2[0-4])\b", h.get_text()) for h in ths):
            main_table = t
            break

    if main_table is None:
        # Sometimes data is in a div-based layout — try <tr> scan
        for t in soup.find_all("table"):
            rows = t.find_all("tr")
            for row in rows:
                cells = row.find_all(["th","td"])
                years_found = [c for c in cells
                               if re.search(r"\b20(1[0-9]|2[0-4])\b", c.get_text())]
                if len(years_found) >= 3:
                    main_table = t
                    break
            if main_table:
                break

    if main_table is None:
        log.debug(f"  {ticker}: no table found in rendered HTML")
        return result

    # Extract year columns from header row
    years            = []
    year_col_indices = []
    for row in main_table.find_all("tr"):
        cells = row.find_all(["th","td"])
        hits  = [(i, int(m.group(1)))
                 for i, c in enumerate(cells)
                 for m in [re.search(r"\b(20\d{2})\b", c.get_text())]
                 if m and 2010 <= int(m.group(1)) <= 2024]
        if len(hits) >= 2:
            years            = [y for _, y in hits]
            year_col_indices = [i for i, _ in hits]
            break

    if not years:
        return result

    # Extract data rows
    found_header = False
    for row in main_table.find_all("tr"):
        cells = row.find_all(["td","th"])

        # Skip until we're past the header row
        if not found_header:
            if any(re.search(r"\b20\d{2}\b", c.get_text()) for c in cells):
                found_header = True
            continue

        if len(cells) < 2:
            continue

        label = norm(cells[0].get_text())
        if not label or len(label) < 2:
            continue

        values = {}

        # Preferred strategy: if the row has one label cell + one value cell per year,
        # align sequentially. This is the typical Biznesradar data-row layout.
        row_value_cells = cells[1:]
        if len(row_value_cells) == len(years):
            for year, cell in zip(years, row_value_cells):
                v = clean_number(cell.get_text())
                if v is not None:
                    values[year] = v
        else:
            # Secondary strategy: map via the indices discovered in the header row.
            for col_idx, year in zip(year_col_indices, years):
                if col_idx < len(cells):
                    v = clean_number(cells[col_idx].get_text())
                    if v is not None:
                        values[year] = v

            # Final fallback: if header-based alignment produced too little, but the row
            # contains at least as many post-label cells as years, try the first N cells
            # after the label sequentially.
            if len(values) <= 1 and len(row_value_cells) >= len(years):
                alt_values = {}
                for year, cell in zip(years, row_value_cells[:len(years)]):
                    v = clean_number(cell.get_text())
                    if v is not None:
                        alt_values[year] = v
                if len(alt_values) > len(values):
                    values = alt_values

        if values:
            result[label] = values

    return result


# ═════════════════════════════════════════════════════════════════════
# SCRAPING
# ═════════════════════════════════════════════════════════════════════

def scrape_firm_selenium(driver, slug: str, ticker: str) -> dict:
    """Scrape all three report types using Selenium."""
    data = {}
    for rtype, url_slug in REPORT_TYPES.items():
        url = f"{BASE_URL}/{url_slug}/{slug},R"
        log.info(f"  {ticker}/{rtype}: {url}")
        try:
            driver.get(url)
            loaded = wait_for_table(driver)
            if not loaded:
                log.warning(f"    Table did not load within {PAGE_TIMEOUT}s — trying without ,R")
                driver.get(f"{BASE_URL}/{url_slug}/{slug}")
                loaded = wait_for_table(driver)
            if not loaded:
                log.warning(f"    Table still not loaded — page may be empty or 404")
                data[rtype] = {}
                continue

            parsed = parse_rows_matrix(extract_table_rows_via_js(driver), ticker)
            if not parsed:
                parsed = parse_rendered_table(driver.page_source, ticker)
            years  = sorted({y for v in parsed.values() for y in v})
            log.info(f"    {len(parsed)} rows, years: {years}")
            data[rtype] = parsed

        except Exception as e:
            log.warning(f"    Selenium error: {e}")
            data[rtype] = {}

        time.sleep(DELAY)
    return data


# ═════════════════════════════════════════════════════════════════════
# FIELD EXTRACTION
# ═════════════════════════════════════════════════════════════════════

FIELD_MAP = {
    # ── Income statement ─────────────────────────────────────────────
    # sale: revenue / przychody ze sprzedazy
    "sale": [
        "przychody ze sprzedazy produktow, towarow i materialow",
        "przychody ze sprzedazy produktow towarow i materialow",
        "przychody netto ze sprzedazy produktow, towarow i materialow",
        "przychody netto ze sprzedazy produktow towarow i materialow",
        "przychody ze sprzedazy i zrownane z nimi",
        "przychody netto ze sprzedazy",
        "przychody ze sprzedazy",
        "przychody ogolem",
    ],
    # cogs: cost of goods sold / koszt wlasny
    "cogs": [
        "koszt wlasny sprzedanych produktow, towarow i materialow",
        "koszt wlasny sprzedazy produktow, towarow i materialow",
        "koszt wlasny sprzedazy produktow towarow i materialow",
        "koszt wlasny sprzedazy",
        "koszty sprzedanych produktow, towarow i materialow",
        "techniczny koszt wytworzenia produkcji sprzedanej",
    ],
    # ni: net income / zysk netto
    "ni": [
        "zysk (strata) netto przypadajacy akcjonariuszom jednostki dominujacej",
        "zysk (strata) netto przypadajacy na akcjonariuszy jednostki dominujacej",
        "zysk netto akcjonariuszy jednostki dominujacej",     # Allegro
        "zysk (strata) netto",
        "zysk netto",
        "wynik netto",
    ],
    # ── Balance sheet ────────────────────────────────────────────────
    # at: total assets
    "at": [
        "aktywa razem",          # most firms (Allegro, Dino, etc.)
        "suma aktywow",
        "aktywa ogolem",
        "suma bilansowa",
        "pasywa razem",          # some firms show it under liabilities side
        "aktywa",
    ],
    # act: current assets
    "act": [
        "aktywa obrotowe razem",
        "razem aktywa obrotowe",
        "aktywa obrotowe",       # Allegro — this exact label
    ],
    # lct: current liabilities
    "lct": [
        "zobowiazania krotkoterminowe razem",
        "razem zobowiazania krotkoterminowe",
        "zobowiazania krotkoterminowe",      # Allegro
        "zobowiazania i rezerwy krotkoterminowe",
        "pasywa obrotowe",
    ],
    # lt: long-term liabilities
    "lt": [
        "zobowiazania dlugoterminowe razem",
        "razem zobowiazania dlugoterminowe",
        "zobowiazania dlugoterminowe",       # Allegro (after norm: dlugoterminowe)
        "zobowiazania i rezerwy dlugoterminowe",
    ],
    # csho: shares outstanding
    # NOTE: Biznesradar does not consistently report shares outstanding in financial tables.
    # If csho is missing, F_EQ_OFFER signal will be NA (treated as 0 in FSCORE sum).
    # This means FSCORE is computed from 8 of 9 signals when csho is unavailable.
    # For the full 9-signal score, retrieve shares from GPW/company annual reports.
    "csho": [
        "liczba akcji (tys.)",
        "liczba akcji w tys",
        "liczba akcji",
        "akcje zwykle (tys.)",
        "akcje zwykle",
    ],
    # ── Cash flow ────────────────────────────────────────────────────
    # cfo: operating cash flow
    "cfo": [
        "przeplywy pieniezne netto z dzialalnosci operacyjnej",
        "srodki pieniezne netto z dzialalnosci operacyjnej",
        "przeplywy pieniezne z dzialalnosci operacyjnej",     # Allegro (after norm)
        "przeplywy z dzialalnosci operacyjnej",
        "i. przeplywy z dzialalnosci operacyjnej",
        "dzialalnos operacyjna",
        "operacyjne przeplywy pieniezne",
    ],
}

# Normalize candidate labels as well, so both sides of matching go through the same transform.
FIELD_MAP = {k: [norm(x) for x in v] for k, v in FIELD_MAP.items()}


def extract_field(
    report: dict,
    candidates: list,
    field_name: str = "",
    ticker: str = "",
    report_name: str = "",
) -> dict:
    """
    Return the best-matching {year: value} dict for a logical field.

    Matching policy:
    1. Collect all substring matches.
    2. Prefer exact equality between candidate and normalized label.
    3. Otherwise prefer the longest candidate (usually the most specific).

    Extra logging is included so mapping failures can be diagnosed directly.
    """
    matches = []

    for cand in candidates:
        for label, vals in report.items():
            if cand in label:
                matches.append((cand, label, vals))

    if not matches:
        log.warning(f"    {ticker}/{report_name}/{field_name}: no match")
        if report:
            log.warning(f"      available labels: {sorted(report.keys())}")
        else:
            log.warning("      report is empty")
        return {}

    # Prefer exact label equality
    for cand, label, vals in matches:
        if cand == label:
            log.info(f"    {ticker}/{report_name}/{field_name}: exact -> {label}")
            return vals

    # Otherwise prefer the longest candidate match (most specific phrase)
    matches.sort(key=lambda x: (len(x[0]), len(x[1])), reverse=True)
    cand, label, vals = matches[0]
    log.info(f"    {ticker}/{report_name}/{field_name}: fuzzy -> {label} (via '{cand}')")
    return vals


def print_labels(ticker: str, scraped: dict) -> None:
    print(f"\n{'='*70}")
    print(f"DEBUG LABELS — {ticker}")
    print(f"{'='*70}")
    for rtype in ("income", "balance", "cashflow"):
        print(f"\n  [{rtype.upper()}]")
        rep = scraped.get(rtype, {})
        if not rep:
            print("    (empty)")
        for label in sorted(rep.keys()):
            sy = sorted(rep[label].keys())[:3]
            sv = [rep[label][y] for y in sy]
            print(f"    {label!r:<65} {dict(zip(sy,sv))}")


def build_panel(ticker: str, scraped: dict, meta: tuple) -> pd.DataFrame:
    _, slug, company, sector, listing_yr, delist_yr, status = meta

    income   = scraped.get("income",   {})
    balance  = scraped.get("balance",  {})
    cashflow = scraped.get("cashflow", {})

    all_years = set()
    for rep in (income, balance, cashflow):
        for v in rep.values():
            all_years |= v.keys()

    start = max(2010, listing_yr or 2010)
    end   = min(2024, delist_yr  or 2024)
    all_years = {y for y in all_years if start <= y <= end}

    if not all_years:
        log.warning(f"  {ticker}: no usable years in [{start}, {end}]")
        return pd.DataFrame()

    fields = {}
    for col, cands in FIELD_MAP.items():
        if col in ("sale", "cogs", "ni"):
            fields[col] = extract_field(income, cands, col, ticker, "income")
        elif col in ("at", "act", "lct", "lt", "csho"):
            fields[col] = extract_field(balance, cands, col, ticker, "balance")
        else:
            fields[col] = extract_field(cashflow, cands, col, ticker, "cashflow")

    for col, vals in fields.items():
        if vals:
            log.info(f"    {ticker}/{col}: years -> {sorted(vals.keys())}")
        else:
            log.warning(f"    {ticker}/{col}: EMPTY")

    rows = []
    for y in sorted(all_years):
        row = dict(ticker=ticker, company_name=company, sector=sector,
                   country="Poland", exchange="WSE",
                   listing_year=listing_yr, status=status,
                   delist_year=delist_yr, year=y)
        for col in ("ni","at","cfo","lt","act","lct","sale","cogs","csho"):
            row[col] = fields[col].get(y, np.nan)
        rows.append(row)

    df      = pd.DataFrame(rows)
    missing = [c for c in ("ni","at","cfo","lt","act","lct","sale","cogs","csho")
               if df[c].isna().all()]
    if missing:
        log.warning(f"  {ticker}: fields still missing: {missing}  "
                    f"→ run --debug-labels {ticker}")
    return df


# ═════════════════════════════════════════════════════════════════════
# F-SCORE
# ═════════════════════════════════════════════════════════════════════

def compute_fscore(df: pd.DataFrame) -> pd.DataFrame:
    chunks = []

    for ticker, g in df.groupby("ticker"):
        g = g.copy().sort_values("year").reset_index(drop=True)

        g["roa"]          = g["ni"]   / g["at"]
        g["cfo_ta"]       = g["cfo"]  / g["at"]
        g["lev"]          = g["lt"]   / g["at"]
        g["curr_ratio"]   = g["act"]  / g["lct"]
        g["gross_margin"] = (g["sale"] - g["cogs"]) / g["sale"]
        g["asset_turn"]   = g["sale"] / g["at"]

        g["F_ROA"]      = (g["roa"] > 0).astype("Int64")
        g["F_CFO"]      = (g["cfo_ta"] > 0).astype("Int64")
        g["F_ACCRUAL"]  = (g["cfo_ta"] > g["roa"]).astype("Int64")
        g["F_DROA"]     = (g["roa"] > g["roa"].shift(1)).astype("Int64")
        g["F_DLEVER"]   = (g["lev"] < g["lev"].shift(1)).astype("Int64")
        g["F_DLIQUID"]  = (g["curr_ratio"] > g["curr_ratio"].shift(1)).astype("Int64")
        g["F_EQ_OFFER"] = (g["csho"] <= g["csho"].shift(1)).astype("Int64")
        g["F_MARGIN"]   = (g["gross_margin"] > g["gross_margin"].shift(1)).astype("Int64")
        g["F_TURN"]     = (g["asset_turn"] > g["asset_turn"].shift(1)).astype("Int64")

        sig = ["F_ROA","F_DROA","F_CFO","F_ACCRUAL",
               "F_DLEVER","F_DLIQUID","F_EQ_OFFER","F_MARGIN","F_TURN"]

        req = {
            "F_ROA": ["ni", "at"],
            "F_CFO": ["cfo", "at"],
            "F_ACCRUAL": ["cfo", "ni", "at"],
            "F_DROA": ["ni", "at"],
            "F_DLEVER": ["lt", "at"],
            "F_DLIQUID": ["act", "lct"],
            "F_EQ_OFFER": ["csho"],
            "F_MARGIN": ["sale", "cogs"],
            "F_TURN": ["sale", "at"],
        }

        for col, needed in req.items():
            missing_now = g[needed].isnull().any(axis=1)
            g.loc[missing_now, col] = pd.NA

        prior_required = {
            "F_DROA": ["ni", "at"],
            "F_DLEVER": ["lt", "at"],
            "F_DLIQUID": ["act", "lct"],
            "F_EQ_OFFER": ["csho"],
            "F_MARGIN": ["sale", "cogs"],
            "F_TURN": ["sale", "at"],
        }

        for col, needed in prior_required.items():
            missing_prev = g[needed].shift(1).isnull().any(axis=1)
            g.loc[missing_prev, col] = pd.NA

        g["N_SIGNALS"] = g[sig].notna().sum(axis=1)
        g["FSCORE"] = g[sig].sum(axis=1, skipna=True).astype("Float64")
        g.loc[g["N_SIGNALS"] == 0, "FSCORE"] = pd.NA

        chunks.append(g)

    return pd.concat(chunks, ignore_index=True)


# ═════════════════════════════════════════════════════════════════════
# ANNUAL RETURNS — Stooq direct CSV
# ═════════════════════════════════════════════════════════════════════

TERMINAL_RETURNS = {
    "GNB.WA": -1.00, "IDA.WA": -0.70,
    "PLM.WA":  0.00, "BOS.WA":  0.00, "CIE.WA": 0.00,
    "ATG.WA":  0.00, "LBW.WA":  0.00, "PGF.WA": 0.00, "TIM.WA": 0.00,
}


def fetch_stooq(ticker_wa: str) -> pd.Series:
    import requests as req
    stooq = ticker_wa.replace(".WA", ".PL").lower()
    url   = f"https://stooq.com/q/d/l/?s={stooq}&d1=20091231&d2=20241231&i=d"
    try:
        r = req.get(url, timeout=20, headers={"User-Agent": "Mozilla/5.0"})
        text = r.text.strip()
        if r.status_code != 200 or not text or len(text) < 20:
            log.warning(f"  {ticker_wa}: Stooq no data")
            return pd.Series(dtype=float)

        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        csv_lines = [ln for ln in lines if "," in ln and re.match(r"^\d{4}-\d{2}-\d{2},", ln)]

        if not csv_lines:
            possible = [ln for ln in lines if "," in ln]
            if len(possible) >= 2:
                csv_lines = possible[1:]

        if not csv_lines:
            log.warning(f"  {ticker_wa}: Stooq returned non-tabular content")
            return pd.Series(dtype=float)

        header = "Date,Open,High,Low,Close,Volume"
        cleaned = "\n".join([header] + csv_lines)
        df = pd.read_csv(io.StringIO(cleaned))

        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df["Close"] = pd.to_numeric(df["Close"], errors="coerce")
        df = df.dropna(subset=["Date", "Close"]).set_index("Date")

        annual = df["Close"].resample("YE").last().dropna()
        log.info(f"  {ticker_wa}: {len(annual)} year-end prices")
        return annual
    except Exception as e:
        log.warning(f"  {ticker_wa}: Stooq error — {e}")
        return pd.Series(dtype=float)

        df       = pd.read_csv(io.StringIO(text))
        date_col = df.columns[0]          # "Data" in Polish
        # Close is 5th column (index 4): Data,Otwarcie,Najwyzszy,Najnizszy,Zamkniecie,Wolumen
        close_col = next(
            (c for c in df.columns if c.lower() in ("zamkniecie","close","zamknięcie")),
            df.columns[4]
        )
        df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
        df = df.dropna(subset=[date_col]).set_index(date_col)
        annual = df[close_col].astype(float).resample("YE").last().dropna()
        log.info(f"  {ticker_wa}: {len(annual)} year-end prices")
        return annual
    except Exception as e:
        log.warning(f"  {ticker_wa}: Stooq error — {e}")
        return pd.Series(dtype=float)


def fetch_annual_returns(tickers: list) -> pd.DataFrame:
    rows = []
    for ticker in tickers:
        prices = fetch_stooq(ticker)
        if prices.empty:
            continue
        for date, ret in prices.pct_change().dropna().items():
            if 2010 <= date.year <= 2024:
                rows.append({"ticker": ticker, "year": date.year, "ret": float(ret)})
        time.sleep(1.5)
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=["ticker","year","ret"])


# ═════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═════════════════════════════════════════════════════════════════════

def write_excel(df: pd.DataFrame, path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl not installed")
        return

    mkfill = lambda h: PatternFill("solid", fgColor=h)
    def mkb():
        s = Side(border_style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    wb = Workbook(); ws = wb.active
    ws.title = "Panel Data"; ws.freeze_panes = "A3"
    cols = list(df.columns); N = len(cols)

    ws.merge_cells(f"A1:{get_column_letter(N)}1")
    t = ws["A1"]
    t.value     = "WSE Piotroski F-Score — Full Panel (Biznesradar.pl + Stooq)"
    t.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill      = mkfill("1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for j, col in enumerate(cols, 1):
        c = ws.cell(row=2, column=j, value=col)
        c.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill      = mkfill("1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = mkb()
        ws.column_dimensions[get_column_letter(j)].width = max(9, min(30, len(col)+2))
    ws.row_dimensions[2].height = 30

    fscore_col = cols.index("FSCORE") + 1 if "FSCORE" in cols else None

    for i, (_, row) in enumerate(df.iterrows()):
        er   = i + 3
        base = "FCE4D6" if row.get("status") == "delisted" else ("F2F2F2" if i%2 else "FFFFFF")
        for j, col in enumerate(cols, 1):
            val = row[col]
            if pd.isna(val): val = None
            cf = base
            if j == fscore_col and val is not None:
                cf = "C6EFCE" if val >= 8 else ("FFC7CE" if val <= 2 else base)
            elif col.startswith("F_") and val is not None:
                cf = "C6EFCE" if int(val) == 1 else "FFC7CE"
            c = ws.cell(row=er, column=j, value=val)
            c.font = Font(name="Arial", size=9)
            c.fill = mkfill(cf); c.border = mkb()
            c.alignment = Alignment(
                horizontal="left" if col in ("company_name","sector") else "center",
                vertical="center")
            if col in ("ni","at","cfo","lt","act","lct","sale","cogs","csho"):
                c.number_format = "#,##0"
            elif col in ("ret","dlret"):
                c.number_format = "0.0%"
            elif col in ("roa","cfo_ta","lev","curr_ratio","gross_margin","asset_turn"):
                c.number_format = "0.000"
        ws.row_dimensions[er].height = 13

    wb.save(path)
    log.info(f"Excel saved: {path}")


# ═════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Biznesradar Selenium scraper")
    parser.add_argument("--no-cache",     action="store_true")
    parser.add_argument("--no-prices",    action="store_true")
    parser.add_argument("--tickers",      nargs="*")
    parser.add_argument("--visible",      action="store_true",
                        help="Show the browser window (not headless)")
    parser.add_argument("--debug-labels", metavar="TICKER")
    parser.add_argument("--debug-mapping", metavar="TICKER",
                        help="Scrape one ticker and run full field-mapping diagnostics")
    parser.add_argument("--delay",        type=float, default=2.0)
    parser.add_argument("--include-banks", action="store_true",
                        help="Include bank tickers in the final panel and scoring")
    parser.add_argument("--include-default-excluded", action="store_true",
                        help="Include tickers that are excluded by default because their pages usually do not load")
    args = parser.parse_args()

    global DELAY
    DELAY = args.delay

    firms = FIRMS
    if not args.include_banks:
        firms = [f for f in firms if f[0] not in BANK_TICKERS]
    if not args.include_default_excluded:
        firms = [f for f in firms if f[0] not in DEFAULT_EXCLUDED_TICKERS]
    if args.tickers:
        requested = {t.upper() for t in args.tickers}
        firms = [f for f in FIRMS if f[0] in requested]
        if not args.include_banks:
            firms = [f for f in firms if f[0] not in BANK_TICKERS]
        if not args.include_default_excluded:
            firms = [f for f in firms if f[0] not in DEFAULT_EXCLUDED_TICKERS]

    if not firms:
        log.error("No firms selected after applying filters.")
        return

    selected_tickers = [f[0] for f in firms]
    excluded_banks = sorted({f[0] for f in FIRMS if f[0] in BANK_TICKERS and f[0] not in selected_tickers})
    excluded_default = sorted({f[0] for f in FIRMS if f[0] in DEFAULT_EXCLUDED_TICKERS and f[0] not in selected_tickers})
    if excluded_banks:
        log.info(f"Excluded bank tickers: {excluded_banks}")
    if excluded_default:
        log.info(f"Excluded default-problem tickers: {excluded_default}")

    log.info("Starting Chrome WebDriver...")
    try:
        driver = make_driver(headless=not args.visible)
    except Exception as e:
        log.error(f"Could not start Chrome: {e}")
        log.error("Make sure Chrome is installed and run: pip install selenium webdriver-manager")
        return

    try:
        # Debug labels mode
        if args.debug_labels:
            target = args.debug_labels.upper()
            match  = next((f for f in FIRMS if f[0] == target), None)
            if not match:
                print(f"Ticker {target} not found"); return
            scraped = scrape_firm_selenium(driver, match[1], target)
            print_labels(target, scraped)
            return

        # Debug mapping mode
        if args.debug_mapping:
            target = args.debug_mapping.upper()
            match  = next((f for f in FIRMS if f[0] == target), None)
            if not match:
                print(f"Ticker {target} not found"); return
            scraped = scrape_firm_selenium(driver, match[1], target)
            print_labels(target, scraped)
            panel = build_panel(target, scraped, match)
            print("\nMAPPED PANEL PREVIEW")
            print("=" * 70)
            if panel.empty:
                print("(empty panel)")
            else:
                cols = ["ticker", "year", "ni", "at", "cfo", "lt", "act", "lct", "sale", "cogs", "csho"]
                print(panel[[c for c in cols if c in panel.columns]].to_string(index=False))
            return

        # ── Scrape ───────────────────────────────────────────────────
        panels = []
        empty_tickers = []
        for meta in firms:
            ticker, slug = meta[0], meta[1]
            cache_path   = CACHE_DIR / f"{ticker.replace('.','_')}.json"

            if cache_path.exists() and not args.no_cache:
                log.info(f"Cache: {ticker}")
                with open(cache_path, encoding="utf-8") as f:
                    raw = json.load(f)
                # Re-normalise all keys (fixes old cache with un-normalised Polish chars)
                scraped = {}
                for rtype, table in raw.items():
                    scraped[rtype] = {norm(k): v for k, v in table.items()}
            else:
                log.info(f"Scraping: {ticker}")
                try:
                    scraped = scrape_firm_selenium(driver, slug, ticker)
                except Exception as e:
                    log.warning(f"Browser error on {ticker}: {e} — restarting browser")
                    try:
                        driver.quit()
                    except Exception:
                        pass
                    time.sleep(5.0)
                    driver = make_driver(headless=not args.visible)
                    try:
                        scraped = scrape_firm_selenium(driver, slug, ticker)
                    except Exception as e2:
                        log.error(f"Retry also failed for {ticker}: {e2} — skipping")
                        scraped = {"income": {}, "balance": {}, "cashflow": {}}
                with open(cache_path, "w", encoding="utf-8") as f:
                    json.dump(scraped, f, indent=2, ensure_ascii=False)
                time.sleep(3.0)

            panel = build_panel(ticker, scraped, meta)
            if not panel.empty:
                panels.append(panel)
            else:
                empty_tickers.append(ticker)
                log.warning(f"{ticker}: empty panel")

        if not panels:
            log.error("No data scraped.")
            return

        if empty_tickers:
            log.warning(f"Tickers dropped due to empty panels: {sorted(empty_tickers)}")

        df = pd.concat(panels, ignore_index=True)
        log.info(f"Panel: {len(df)} rows, {df['ticker'].nunique()} firms")

        # ── F-Score ──────────────────────────────────────────────────
        df = compute_fscore(df)
        n_scored = int(df["FSCORE"].notna().sum())
        log.info(f"F-Score: {n_scored}/{len(df)} rows scored")

        # ── Returns ──────────────────────────────────────────────────
        df["ret"] = np.nan; df["dlret"] = np.nan
        if not args.no_prices:
            log.info("Fetching Stooq returns...")
            rets = fetch_annual_returns([f[0] for f in firms])
            if not rets.empty:
                df = df.merge(rets, on=["ticker","year"], how="left", suffixes=("","_s"))
                if "ret_s" in df.columns:
                    df["ret"] = df["ret_s"].combine_first(df["ret"])
                    df.drop(columns=["ret_s"], inplace=True)

        for ticker, dlret in TERMINAL_RETURNS.items():
            m = df["ticker"] == ticker
            if not m.any(): continue
            dy = df.loc[m,"delist_year"].iloc[0]
            if pd.notna(dy):
                df.loc[m & (df["year"]==int(dy)), "dlret"] = dlret

        # ── Column order & save ───────────────────────────────────────
        col_order = [
            "ticker","company_name","sector","country","exchange",
            "listing_year","status","delist_year","year",
            "ni","at","cfo","lt","act","lct","sale","cogs","csho",
            "roa","cfo_ta","lev","curr_ratio","gross_margin","asset_turn",
            "F_ROA","F_DROA","F_CFO","F_ACCRUAL",
            "F_DLEVER","F_DLIQUID","F_EQ_OFFER","F_MARGIN","F_TURN","N_SIGNALS","FSCORE",
            "ret","dlret",
        ]
        df = df[[c for c in col_order if c in df.columns]]
        df = df.sort_values(["ticker","year"]).reset_index(drop=True)

        fin = ["ticker","company_name","year","ni","at","cfo","lt","act","lct","sale","cogs","csho"]
        df[[c for c in fin if c in df.columns]].to_csv(OUT_DIR/"financials_raw_production.csv", index=False)
        df.to_csv(OUT_DIR/"wse_piotroski_production.csv", index=False)
        write_excel(df, str(OUT_DIR/"wse_piotroski_production.xlsx"))

        # ── Summary ───────────────────────────────────────────────────
        fin_cols = ["ni","at","cfo","lt","act","lct","sale","cogs","csho"]
        print(f"\n{'='*75}")
        print("COMPLETE")
        print(f"{'='*75}")
        print(f"{'Ticker':<10} {'Status':<10} {'Yrs':<5} {'Scored':<7} {'F-Score':<10} Missing")
        print(f"{'-'*75}")
        for ticker, g in df.groupby("ticker"):
            n_yrs   = len(g)
            n_score = int(g["FSCORE"].notna().sum())
            rng     = f"{int(g['FSCORE'].min())}–{int(g['FSCORE'].max())}" if n_score else "N/A"
            miss    = [c for c in fin_cols if g[c].isna().all()]
            print(f"{ticker:<10} {g['status'].iloc[0]:<10} {n_yrs:<5} {n_score:<7} "
                  f"{rng:<10} {','.join(miss) or 'none'}")
        print(f"{'='*75}")
        if excluded_banks:
            print(f"Excluded banks: {', '.join(excluded_banks)}")
        if excluded_default:
            print(f"Default excluded tickers: {', '.join(excluded_default)}")
        if 'empty_tickers' in locals() and empty_tickers:
            print(f"Dropped empty panels: {', '.join(sorted(empty_tickers))}")
        print(f"\nOutputs: {OUT_DIR.resolve()}/")

    finally:
        driver.quit()
        log.info("Browser closed.")


if __name__ == "__main__":
    main()
